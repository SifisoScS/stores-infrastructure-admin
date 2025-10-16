#!/usr/bin/env python3
"""
Safe script to add sample data to existing Excel file without overwriting structure
"""
import pandas as pd
from openpyxl import load_workbook

def add_sample_data_safely():
    """Add sample data to existing sheets without overwriting the file structure"""
    
    excel_file = 'STORES_INFRASTRUCTURE_ADMINISTRATION.xlsx'
    
    # Sample data for just a few categories to demonstrate
    sample_data = {
        'Electric': [
            ['ELE001', 'LED Light Bulbs 60W', 25, 'pieces', 'Store A', 50, 100, 'ElectroMax Ltd', '2025-07-15', 'N/A', 15.50, 387.50],
            ['ELE002', 'Extension Cords 5m', 8, 'pieces', 'Store B', 15, 50, 'PowerPlus Co', '2025-08-01', '2026-08-01', 85.00, 680.00],
            ['ELE003', 'Circuit Breakers 20A', 12, 'pieces', 'Store A', 20, 60, 'SafeElectric Inc', '2025-07-20', 'N/A', 125.00, 1500.00]
        ],
        'Plumbing': [
            ['PLM001', 'PVC Pipes 110mm', 15, 'meters', 'Store A', 25, 80, 'AquaFlow Systems', '2025-08-05', 'N/A', 45.00, 675.00],
            ['PLM002', 'Toilet Seats Standard', 6, 'pieces', 'Store B', 10, 25, 'BathCare Ltd', '2025-07-25', 'N/A', 185.50, 1113.00]
        ],
        'Safety': [
            ['SAF001', 'Safety Helmets', 18, 'pieces', 'Store A', 25, 60, 'SafeGuard Ltd', '2025-08-10', 'N/A', 125.00, 2250.00],
            ['SAF002', 'Safety Vests Hi-Vis', 8, 'pieces', 'Store B', 15, 40, 'VisProtect SA', '2025-07-30', 'N/A', 85.00, 680.00]
        ]
    }
    
    # Load the existing workbook
    book = load_workbook(excel_file)
    
    # Add sample data to specific sheets
    for sheet_name, data in sample_data.items():
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            
            # Add data starting from row 2 (row 1 has headers)
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            print(f"+ Added {len(data)} items to {sheet_name}")
    
    # Add some maintenance log entries
    if 'Maintenance Log' in book.sheetnames:
        maintenance_data = [
            ['2025-08-15', 'Replaced LED bulbs in Store A', 'Electric', 'John Smith', '2h', 'Easy replacement', '', 'Improved lighting'],
            ['2025-08-20', 'Fixed leaking pipe in Store B', 'Plumbing', 'Sarah Connor', '1.5h', 'Minor leak repair', '', 'Prevented water damage'],
            ['2025-08-25', 'Safety inspection all stores', 'Safety', 'Mike Wilson', '3h', 'Annual safety check', '', 'All equipment certified']
        ]
        
        sheet = book['Maintenance Log']
        for row_idx, row_data in enumerate(maintenance_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"+ Added {len(maintenance_data)} maintenance entries")
    
    # Add some suppliers
    if 'Suppliers & Contractors' in book.sheetnames:
        suppliers_data = [
            ['ElectroMax Ltd', 'Electric', 'David Johnson', 'david@electromax.co.za / 011-555-0001', '2025-12-31', 'Preferred'],
            ['AquaFlow Systems', 'Plumbing', 'Maria Rodriguez', 'maria@aquaflow.co.za / 021-555-0002', '2026-06-30', 'Approved'],
            ['SafeGuard Ltd', 'Safety', 'Susan Williams', 'susan@safeguard.co.za / 031-555-0006', '2026-01-31', 'Preferred']
        ]
        
        sheet = book['Suppliers & Contractors']
        for row_idx, row_data in enumerate(suppliers_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"+ Added {len(suppliers_data)} suppliers")
    
    # Save the workbook
    book.save(excel_file)
    book.close()
    
    print(f"\nSUCCESS: Sample data added to {excel_file}!")
    print("\nSample data added:")
    print("- Electric: 3 items (1 low stock)")
    print("- Plumbing: 2 items (1 low stock)")
    print("- Safety: 2 items (1 low stock)")
    print("- Maintenance Log: 3 recent activities")
    print("- Suppliers: 3 supplier contacts")
    print("\nNow click 'Reload Data' in the web application to see the changes!")

if __name__ == "__main__":
    try:
        add_sample_data_safely()
    except Exception as e:
        print(f"ERROR: {e}")
        print("Make sure the Excel file is closed before running this script.")