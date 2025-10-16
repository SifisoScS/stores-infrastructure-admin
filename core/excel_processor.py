#!/usr/bin/env python3
"""
Advanced Excel Processing Engine for Derivco Facilities Management System
Handles all Excel file operations with validation, error handling, and automation
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional, Tuple, Any, Union
import os
from datetime import datetime, timedelta
import logging
from dataclasses import dataclass
import numpy as np
from pathlib import Path
import json

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ExcelFileConfig:
    """Configuration for Excel file processing"""
    filename: str
    sheets: List[str]
    required_columns: Dict[str, List[str]]
    data_types: Dict[str, str]
    validation_rules: Dict[str, Any]

class ExcelProcessingError(Exception):
    """Custom exception for Excel processing errors"""
    pass

class AdvancedExcelProcessor:
    """
    Comprehensive Excel processing engine for all facility management data
    Handles inventory, sign-out, medical, and reporting operations
    """

    def __init__(self, base_path: str = "."):
        self.base_path = Path(base_path)
        self.config = self._load_excel_configs()
        self.cache = {}
        self.validation_errors = []

    def _load_excel_configs(self) -> Dict[str, ExcelFileConfig]:
        """Load Excel file configurations"""
        return {
            'inventory': ExcelFileConfig(
                filename='STORES_INFRASTRUCTURE_ADMINISTRATION_enhanced.xlsx',
                sheets=['Electric', 'Plumbing', 'Carpentry', 'Painting', 'Aircon', 'Ceiling Tiles',
                       'Decoration', 'Parking & Signage', 'Safety', 'Access Control'],
                required_columns={
                    'default': ['Item Code', 'Item Name', 'Quantity', 'Unit Price', 'Total Value', 'Supplier']
                },
                data_types={
                    'Quantity': 'int64',
                    'Unit Price': 'float64',
                    'Total Value': 'float64'
                },
                validation_rules={
                    'min_quantity': 0,
                    'max_quantity': 10000,
                    'min_price': 0,
                    'max_price': 50000
                }
            ),
            'signout': ExcelFileConfig(
                filename='signout_data_improved.xlsx',
                sheets=['Sheet1'],
                required_columns={
                    'default': ['Item_Name', 'Borrower_Name', 'WO_REQ_No', 'Project_Type',
                               'Department', 'Date_Out', 'Time_Out', 'Expected_Return', 'Date_In', 'Time_In']
                },
                data_types={
                    'Date_Out': 'datetime64[ns]',
                    'Date_In': 'datetime64[ns]'
                },
                validation_rules={
                    'required_fields': ['Item_Name', 'Borrower_Name', 'Date_Out']
                }
            ),
            'medical': ExcelFileConfig(
                filename='medication_data_enhanced.xlsx',
                sheets=['Medical_Inventory', 'Patient_Records'],
                required_columns={
                    'Medical_Inventory': ['Item_Code', 'Item_Name', 'Category', 'Stock_Level', 'Expiry_Date'],
                    'Patient_Records': ['Date', 'Patient_Name', 'Incident_Type', 'Treatment', 'Status']
                },
                data_types={
                    'Stock_Level': 'int64',
                    'Expiry_Date': 'datetime64[ns]'
                },
                validation_rules={
                    'expiry_warning_days': 30,
                    'min_stock_level': 0
                }
            )
        }

    def read_excel_file(self, file_type: str, sheet_name: str = None,
                       use_cache: bool = True) -> pd.DataFrame:
        """
        Read Excel file with caching and error handling

        Args:
            file_type: Type of Excel file ('inventory', 'signout', 'medical')
            sheet_name: Specific sheet to read (optional)
            use_cache: Whether to use cached data

        Returns:
            pandas DataFrame with the data
        """
        try:
            if file_type not in self.config:
                raise ExcelProcessingError(f"Unknown file type: {file_type}")

            config = self.config[file_type]
            file_path = self.base_path / config.filename

            if not file_path.exists():
                raise ExcelProcessingError(f"Excel file not found: {file_path}")

            cache_key = f"{file_type}_{sheet_name or 'all'}"

            if use_cache and cache_key in self.cache:
                logger.info(f"Using cached data for {cache_key}")
                return self.cache[cache_key]

            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                # Read all sheets and combine
                all_sheets = pd.read_excel(file_path, sheet_name=None)
                df = pd.concat(all_sheets.values(), ignore_index=True)

            # Apply data type conversions
            for col, dtype in config.data_types.items():
                if col in df.columns:
                    try:
                        if dtype.startswith('datetime'):
                            df[col] = pd.to_datetime(df[col], errors='coerce')
                        else:
                            df[col] = df[col].astype(dtype, errors='ignore')
                    except Exception as e:
                        logger.warning(f"Could not convert {col} to {dtype}: {e}")

            # Cache the result
            if use_cache:
                self.cache[cache_key] = df

            logger.info(f"Successfully loaded {len(df)} records from {file_path}")
            return df

        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise ExcelProcessingError(f"Failed to read {file_type} data: {e}")

    def validate_data(self, df: pd.DataFrame, file_type: str) -> Tuple[bool, List[str]]:
        """
        Validate Excel data against business rules

        Args:
            df: DataFrame to validate
            file_type: Type of data being validated

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []
        config = self.config.get(file_type, {})
        rules = config.validation_rules if hasattr(config, 'validation_rules') else {}

        try:
            # Check required columns
            required_cols = config.required_columns.get('default', [])
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                errors.append(f"Missing required columns: {missing_cols}")

            # File-specific validations
            if file_type == 'inventory':
                # Validate inventory data
                if 'Quantity' in df.columns:
                    invalid_qty = df[(df['Quantity'] < 0) | (df['Quantity'] > rules.get('max_quantity', 10000))]
                    if not invalid_qty.empty:
                        errors.append(f"Invalid quantities found in {len(invalid_qty)} records")

                if 'Unit Price' in df.columns:
                    invalid_price = df[(df['Unit Price'] < 0) | (df['Unit Price'] > rules.get('max_price', 50000))]
                    if not invalid_price.empty:
                        errors.append(f"Invalid prices found in {len(invalid_price)} records")

            elif file_type == 'signout':
                # Validate sign-out data
                required_fields = rules.get('required_fields', [])
                for field in required_fields:
                    if field in df.columns:
                        missing_data = df[df[field].isna() | (df[field] == '')]
                        if not missing_data.empty:
                            errors.append(f"Missing {field} in {len(missing_data)} records")

            elif file_type == 'medical':
                # Validate medical data
                if 'Expiry_Date' in df.columns:
                    today = datetime.now()
                    expired_items = df[df['Expiry_Date'] < today]
                    if not expired_items.empty:
                        errors.append(f"Found {len(expired_items)} expired medical items")

                if 'Stock_Level' in df.columns:
                    negative_stock = df[df['Stock_Level'] < 0]
                    if not negative_stock.empty:
                        errors.append(f"Found {len(negative_stock)} items with negative stock")

            is_valid = len(errors) == 0
            self.validation_errors = errors

            return is_valid, errors

        except Exception as e:
            logger.error(f"Error validating data: {e}")
            return False, [f"Validation error: {e}"]

    def write_excel_file(self, data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
                        file_path: str, file_type: str = 'custom',
                        apply_formatting: bool = True) -> bool:
        """
        Write data to Excel file with professional formatting

        Args:
            data: DataFrame or dict of DataFrames to write
            file_path: Output file path
            file_type: Type of file for formatting rules
            apply_formatting: Whether to apply professional formatting

        Returns:
            Success status
        """
        try:
            output_path = self.base_path / file_path

            if isinstance(data, pd.DataFrame):
                # Single sheet
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    data.to_excel(writer, sheet_name='Data', index=False)

                    if apply_formatting:
                        self._apply_formatting(writer.book['Data'], data)
            else:
                # Multiple sheets
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for sheet_name, df in data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        if apply_formatting:
                            self._apply_formatting(writer.book[sheet_name], df)

            logger.info(f"Successfully wrote Excel file: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            return False

    def _apply_formatting(self, worksheet, df: pd.DataFrame):
        """Apply professional formatting to Excel worksheet"""
        try:
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col_num, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Number formatting for financial columns
            for col_num, col_name in enumerate(df.columns, 1):
                if any(keyword in col_name.lower() for keyword in ['price', 'value', 'cost']):
                    for row_num in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.number_format = 'R#,##0.00'

        except Exception as e:
            logger.warning(f"Error applying formatting: {e}")

    def generate_inventory_report(self, include_analytics: bool = True) -> Dict[str, Any]:
        """
        Generate comprehensive inventory report

        Args:
            include_analytics: Whether to include advanced analytics

        Returns:
            Dictionary containing report data and analytics
        """
        try:
            report = {
                'timestamp': datetime.now().isoformat(),
                'summary': {},
                'categories': {},
                'alerts': [],
                'analytics': {}
            }

            # Read all inventory sheets
            config = self.config['inventory']
            all_data = []

            for sheet in config.sheets:
                try:
                    df = self.read_excel_file('inventory', sheet)
                    df['Category'] = sheet
                    all_data.append(df)
                except Exception as e:
                    logger.warning(f"Could not read sheet {sheet}: {e}")

            if not all_data:
                return report

            combined_df = pd.concat(all_data, ignore_index=True)

            # Generate summary statistics
            report['summary'] = {
                'total_items': len(combined_df),
                'total_categories': len(config.sheets),
                'total_value': combined_df['Total Value'].sum() if 'Total Value' in combined_df.columns else 0,
                'low_stock_items': len(combined_df[combined_df['Quantity'] < 10]) if 'Quantity' in combined_df.columns else 0
            }

            # Category analysis
            for category in config.sheets:
                category_data = combined_df[combined_df['Category'] == category]
                if not category_data.empty:
                    report['categories'][category] = {
                        'item_count': len(category_data),
                        'total_value': category_data['Total Value'].sum() if 'Total Value' in category_data.columns else 0,
                        'avg_price': category_data['Unit Price'].mean() if 'Unit Price' in category_data.columns else 0,
                        'low_stock_items': len(category_data[category_data['Quantity'] < 10]) if 'Quantity' in category_data.columns else 0
                    }

            # Generate alerts
            if 'Quantity' in combined_df.columns:
                low_stock = combined_df[combined_df['Quantity'] < 5]
                for _, item in low_stock.iterrows():
                    report['alerts'].append({
                        'type': 'low_stock',
                        'severity': 'high' if item['Quantity'] < 2 else 'medium',
                        'item': item['Item Name'] if 'Item Name' in item else 'Unknown',
                        'current_stock': item['Quantity'],
                        'category': item['Category']
                    })

            # Advanced analytics
            if include_analytics:
                report['analytics'] = self._generate_inventory_analytics(combined_df)

            return report

        except Exception as e:
            logger.error(f"Error generating inventory report: {e}")
            return {'error': str(e)}

    def _generate_inventory_analytics(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate advanced inventory analytics"""
        analytics = {}

        try:
            # Value distribution analysis
            if 'Total Value' in df.columns:
                analytics['value_distribution'] = {
                    'mean': float(df['Total Value'].mean()),
                    'median': float(df['Total Value'].median()),
                    'std': float(df['Total Value'].std()),
                    'quartiles': df['Total Value'].quantile([0.25, 0.5, 0.75]).to_dict()
                }

            # Stock level analysis
            if 'Quantity' in df.columns:
                analytics['stock_analysis'] = {
                    'total_items': int(df['Quantity'].sum()),
                    'avg_stock_per_item': float(df['Quantity'].mean()),
                    'items_needing_reorder': int(len(df[df['Quantity'] < 10])),
                    'overstocked_items': int(len(df[df['Quantity'] > 100]))
                }

            # Supplier analysis
            if 'Supplier' in df.columns:
                supplier_counts = df['Supplier'].value_counts().head(10)
                analytics['top_suppliers'] = supplier_counts.to_dict()

            # Category performance
            if 'Category' in df.columns:
                category_stats = df.groupby('Category').agg({
                    'Quantity': 'sum',
                    'Total Value': 'sum' if 'Total Value' in df.columns else 'count'
                }).to_dict('index')
                analytics['category_performance'] = category_stats

        except Exception as e:
            logger.warning(f"Error in analytics generation: {e}")

        return analytics

    def generate_compliance_report(self) -> Dict[str, Any]:
        """
        Generate facilities compliance report based on sign-out data

        Returns:
            Comprehensive compliance analysis
        """
        try:
            signout_df = self.read_excel_file('signout')

            report = {
                'timestamp': datetime.now().isoformat(),
                'compliance_metrics': {},
                'violations': [],
                'recommendations': []
            }

            # Work Order compliance analysis
            if 'WO_REQ_No' in signout_df.columns:
                # Define supervisor names that shouldn't be work orders
                supervisor_names = {
                    'Erasmus Ngwane', 'Brindley Harrington', 'Not Provided', 'Not provided',
                    'not provided', 'N/A', 'n/a', 'None', 'none', ''
                }

                missing_wo = signout_df[
                    signout_df['WO_REQ_No'].isna() |
                    signout_df['WO_REQ_No'].isin(supervisor_names)
                ]

                report['compliance_metrics']['missing_work_orders'] = {
                    'count': len(missing_wo),
                    'percentage': (len(missing_wo) / len(signout_df)) * 100,
                    'total_transactions': len(signout_df)
                }

            # Tool return compliance
            if 'Date_In' in signout_df.columns:
                unreturned = signout_df[signout_df['Date_In'].isna()]
                report['compliance_metrics']['unreturned_tools'] = {
                    'count': len(unreturned),
                    'percentage': (len(unreturned) / len(signout_df)) * 100,
                    'estimated_value': len(unreturned) * 500  # Estimated R500 per tool
                }

            # Generate specific violations
            for _, violation in missing_wo.iterrows():
                report['violations'].append({
                    'type': 'missing_work_order',
                    'item': violation.get('Item_Name', 'Unknown'),
                    'borrower': violation.get('Borrower_Name', 'Unknown'),
                    'date': violation.get('Date_Out', 'Unknown')
                })

            # Generate recommendations
            if report['compliance_metrics'].get('missing_work_orders', {}).get('percentage', 0) > 50:
                report['recommendations'].append({
                    'type': 'policy_enforcement',
                    'priority': 'high',
                    'description': 'Implement mandatory work order validation before tool sign-out'
                })

            return report

        except Exception as e:
            logger.error(f"Error generating compliance report: {e}")
            return {'error': str(e)}

    def backup_excel_files(self, backup_dir: str = 'backups') -> bool:
        """Create timestamped backups of all Excel files"""
        try:
            backup_path = self.base_path / backup_dir
            backup_path.mkdir(exist_ok=True)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            for file_type, config in self.config.items():
                source_file = self.base_path / config.filename
                if source_file.exists():
                    backup_file = backup_path / f"{file_type}_{timestamp}_{config.filename}"
                    source_file.replace(backup_file)
                    logger.info(f"Backed up {source_file} to {backup_file}")

            return True

        except Exception as e:
            logger.error(f"Error creating backups: {e}")
            return False

    def clear_cache(self):
        """Clear the data cache"""
        self.cache.clear()
        logger.info("Data cache cleared")

    def get_cache_info(self) -> Dict[str, Any]:
        """Get information about cached data"""
        return {
            'cached_datasets': list(self.cache.keys()),
            'cache_size': len(self.cache),
            'memory_usage': sum(df.memory_usage(deep=True).sum() for df in self.cache.values())
        }